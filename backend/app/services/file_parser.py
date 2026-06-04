from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import pdfplumber


@dataclass(slots=True)
class ParsedTabularFile:
	file_type: str
	sheet_names: list[str]
	selected_source: str
	columns: list[str]
	rows: list[dict[str, Any]]

	@property
	def preview_rows(self) -> list[dict[str, Any]]:
		return self.rows[:5]


def extract_tabular_file(file_path: Path, sheet_name: str | None = None) -> ParsedTabularFile:
	suffix = file_path.suffix.lower()
	if suffix == ".xlsx":
		return _extract_xlsx(file_path, sheet_name=sheet_name)
	if suffix == ".pdf":
		return _extract_pdf(file_path)
	raise ValueError("Only PDF and XLSX files are supported")


def _extract_xlsx(file_path: Path, sheet_name: str | None = None) -> ParsedTabularFile:
	workbook = load_workbook(file_path, read_only=True, data_only=True)
	try:
		sheet_names = list(workbook.sheetnames)
		if not sheet_names:
			raise ValueError("The XLSX file does not contain any sheets")

		target_sheet_name = sheet_name or sheet_names[0]
		if target_sheet_name not in workbook.sheetnames:
			raise ValueError(f"Sheet '{target_sheet_name}' was not found in the workbook")

		sheet = workbook[target_sheet_name]
		rows_iter = sheet.iter_rows(values_only=True)
		headers, data_rows = _extract_rows_from_iterable(rows_iter)
		return ParsedTabularFile(
			file_type="xlsx",
			sheet_names=sheet_names,
			selected_source=target_sheet_name,
			columns=headers,
			rows=data_rows,
		)
	finally:
		workbook.close()


def _extract_pdf(file_path: Path) -> ParsedTabularFile:
	with pdfplumber.open(file_path) as pdf:
		for page_index, page in enumerate(pdf.pages, start=1):
			tables = page.extract_tables() or []
			for table_index, table in enumerate(tables, start=1):
				headers, data_rows = _extract_rows_from_table(table)
				if headers and data_rows:
					return ParsedTabularFile(
						file_type="pdf",
						sheet_names=[],
						selected_source=f"Page {page_index}, Table {table_index}",
						columns=headers,
						rows=data_rows,
					)

			text = page.extract_text() or ""
			headers, data_rows = _extract_rows_from_text(text)
			if headers and data_rows:
				return ParsedTabularFile(
					file_type="pdf",
					sheet_names=[],
					selected_source=f"Page {page_index}",
					columns=headers,
					rows=data_rows,
				)

	raise ValueError("Could not extract a readable table structure from the PDF file")


def _extract_rows_from_iterable(rows_iter: Any) -> tuple[list[str], list[dict[str, Any]]]:
	for raw_row in rows_iter:
		if _row_has_values(raw_row):
			headers = _deduplicate_headers(list(raw_row))
			data_rows: list[dict[str, Any]] = []
			for data_row in rows_iter:
				if not _row_has_values(data_row):
					continue
				data_rows.append(_row_to_dict(headers, data_row))
			return headers, data_rows
	return [], []


def _extract_rows_from_table(table: list[list[Any]] | None) -> tuple[list[str], list[dict[str, Any]]]:
	if not table:
		return [], []
	for index, raw_row in enumerate(table):
		if _row_has_values(raw_row):
			headers = _deduplicate_headers(list(raw_row))
			data_rows = [
				_row_to_dict(headers, data_row)
				for data_row in table[index + 1 :]
				if _row_has_values(data_row)
			]
			return headers, data_rows
	return [], []


def _extract_rows_from_text(text: str) -> tuple[list[str], list[dict[str, Any]]]:
	lines = [line.strip() for line in text.splitlines() if line.strip()]
	if len(lines) < 2:
		return [], []

	headers = _split_line(lines[0])
	if not headers:
		return [], []
	headers = _deduplicate_headers(headers)
	data_rows = []
	for line in lines[1:]:
		parts = _split_line(line)
		if not parts:
			continue
		data_rows.append(_row_to_dict(headers, parts))
	return headers, data_rows


def _split_line(line: str) -> list[str]:
	parts = [segment.strip() for segment in re.split(r"\t+|\s{2,}", line) if segment.strip()]
	if len(parts) <= 1:
		parts = [segment.strip() for segment in line.split("|") if segment.strip()]
	return parts


def _row_has_values(row: Any) -> bool:
	return any(cell not in (None, "") for cell in row or [])


def _deduplicate_headers(raw_headers: list[Any]) -> list[str]:
	counts: dict[str, int] = {}
	labels: list[str] = []
	for index, value in enumerate(raw_headers, start=1):
		label = str(value).strip() if value not in (None, "") else f"Column {index}"
		occurrences = counts.get(label, 0)
		counts[label] = occurrences + 1
		if occurrences:
			labels.append(f"{label} ({occurrences + 1})")
		else:
			labels.append(label)
	return labels


def _row_to_dict(headers: list[str], row: Any) -> dict[str, Any]:
	row_dict: dict[str, Any] = {}
	for index, header in enumerate(headers):
		value = row[index] if index < len(row) else None
		row_dict[header] = _json_safe_value(value)
	return row_dict


def _json_safe_value(value: Any) -> Any:
	if isinstance(value, Decimal):
		return float(value)
	if hasattr(value, "isoformat") and callable(value.isoformat):
		return value.isoformat()
	if isinstance(value, bytes):
		return value.decode("utf-8", errors="ignore")
	return value


def to_decimal(value: Any) -> Decimal:
	if value is None:
		return Decimal("0")
	if isinstance(value, Decimal):
		return value
	if isinstance(value, bool):
		return Decimal(int(value))
	if isinstance(value, (int, float)):
		return Decimal(str(value))
	text = str(value).strip().replace(",", "")
	if not text:
		return Decimal("0")
	try:
		return Decimal(text)
	except (InvalidOperation, ValueError):
		return Decimal("0")


def to_text(value: Any) -> str | None:
	if value is None:
		return None
	text = str(value).strip()
	return text or None